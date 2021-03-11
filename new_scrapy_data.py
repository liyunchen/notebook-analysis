# -*- coding: utf-8 -*-


"""
李运辰 2021-3-11

公众号：python爬虫数据分析挖掘
"""


import requests
from lxml import etree
import json
import time


headers = {
            'user-agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3947.100 Safari/537.36',
        }


import openpyxl
outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)
outws.cell(row=1, column=1, value="index")
outws.cell(row=1, column=2, value="brand")
outws.cell(row=1, column=3, value="title")
outws.cell(row=1, column=4, value="price")
count = 2


###获取每一页的商品数据
def getlist(url,brand):
    global  count
    #url="https://search.jd.com/search?keyword=笔记本&wq=笔记本&ev=exbrand_联想%5E&page=9&s=241&click=1"
    res = requests.get(url,headers=headers)
    res.encoding = 'utf-8'
    text = res.text


    selector = etree.HTML(text)
    list = selector.xpath('//*[@id="J_goodsList"]/ul/li')

    for i in list:
        title=i.xpath('.//div[@class="p-name p-name-type-2"]/a/em/text()')[0]
        price = i.xpath('.//div[@class="p-price"]/strong/i/text()')[0]
        #product_id = i.xpath('.//div[@class="p-commit"]/strong/a/@id')[0].replace("J_comment_","")


        ###获取商品评论数
        #comment_count = commentcount(product_id)
        #print("title="+str(title))
        #print("price="+str(price))
        #print("comment_count="+str(comment_count))

        outws.cell(row=count, column=1, value=str(count-1))
        outws.cell(row=count, column=2, value=str(brand))
        outws.cell(row=count, column=3, value=str(title))
        outws.cell(row=count, column=4, value=str(price))


        count = count +1
        #print("目前条数="+str(count))
        #print("-----")



#遍历每一页
def getpage(brand_dict):
    global  count
    for k, v in brand_dict.items():

        page = 1
        s = 1
        brand = str(k)
        try:
            for i in range(1, int(v) + 1):
                url = "https://search.jd.com/search?keyword=笔记本&wq=笔记本&ev=exbrand_" + str(brand) + "&page=" + str(
                    page) + "&s=" + str(s) + "&click=1"
                getlist(url, brand)
                page = page + 2
                s = s + 60
                # if int(i) %3==0:
                #     time.sleep(2)
                print("品牌=" + str(k) + ",页数=" + str(v) + ",当前页数=" + str(i))
        except:
            pass


    outwb.save("笔记本电脑-李运辰0311.xls")  # 保存

#开始爬取
brand_dict={
    '联想（lenovo）':100,
    'ThinkPad':100,
    '戴尔（DELL）':100,
    '惠普（HP）':100,
    '华为（HUAWEI）':100,
    'Apple':100,
    '小米（MI）':47,
    '宏碁（acer）':43,
    '荣耀（HONOR）':21,
    '机械革命（MECHREVO）':31,
    '微软（Microsoft）':100,
    'LG':3,
    '神舟（HASEE）':34,
    'VAIO':3,
    '三星（SAMSUNG）':47,
}

getpage(brand_dict)